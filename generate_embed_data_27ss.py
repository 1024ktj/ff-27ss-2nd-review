import sys
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
import json
import os
import glob
import shutil
from datetime import datetime, date
from collections import Counter

# ─── Paths ────────────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# 가장 최근 수정일자의 엑셀 파일 자동 선택 (OneDrive 원본 직접 참조)
_candidates = glob.glob(r"C:\Users\AD1082\OneDrive - F&F\바탕 화면\(WEAR)27SS TRADE SHOW LIST*.xlsx")
if not _candidates:
    raise FileNotFoundError("(WEAR)27SS TRADE SHOW LIST*.xlsx 파일을 찾을 수 없습니다.")
EXCEL_PATH = max(_candidates, key=os.path.getmtime)
print(f"📂 사용 엑셀: {os.path.basename(EXCEL_PATH)}")

OUTPUT_PATH = os.path.join(BASE_DIR, "embed_data.js")
TMP_PATH    = r"C:\Users\AD1082\AppData\Local\Temp\trade_tmp.xlsx"

# ─── Vendor KR → EN ───────────────────────────────────────────────────────────
VENDOR_KR_TO_EN = {
    '(주) 약진통상':         'YAKJIN TRADING',
    '(주)기도산업':           'KIDO INDUSTRIAL CO. LTD.,',
    '(주)다인지아이씨':       'Dain.g.i.c',
    '(주)팬코':               'PANKO',
    '(주)포마트코퍼레이션':   'Foremart corporation',
    '㈜노브랜드':             'NOBLAND INTERNATIONAL INC',
    '㈜노브랜드(우븐)':       'NOBLAND INTERNATIONAL INC',
    '원전교역':               'WONJEON CORPORATION',
    '주식회사 거림씨앤에프':  'GEU LIM CNF CO.,LTD',
    '티피나디아㈜':           'TP Nadia Co.,Ltd',
    '한솔섬유 ㈜':            'HANSOLL TEXTILE LTD.',
    '한솔섬유 (주)':          'HANSOLL TEXTILE LTD.',
}

# ─── Transit days by CO ───────────────────────────────────────────────────────
TRANSIT_BY_CO = {'CN': 2, 'VN': 5, 'KR': 0}

# ─── Column indices (0-based, 1-indexed column letter - 1) ───────────────────
#
#  Excel  Letter  0-based   Field
#  ─────  ──────  ───────   ─────
#  2      B       1         main_pbn
#  3      C       2         sku
#  4      D       3         sort
#  5      E       4         class_
#  6      F       5         item_type
#  8      H       7         style_code
#  9      I       8         color
#  10     J       9         description
#  12     L       11        go_drop
#  15     O       14        designer
#  16     P       15        sourcing
#  17     Q       16        vendor
#  20     T       19        cn_show_s
#  21     U       20        cn_show_l
#  22     V       21        cn_show_xl    ← new
#  23     W       22        cn_show_pcs   (TOTAL)
#  24     X       23        cn_focus_s
#  25     Y       24        cn_focus_m    ← new
#  26     Z       25        cn_focus_xl
#  27     AA      26        cn_focus_pcs  (TOTAL)
#  30     AD      29        kr_keep_pcs   (KEEP KOREA TOTAL)
#  35     AI      34        ttl_pcs       (TTL TOTAL)
#  36     AJ      35        kr_expected
#  39     AM      38        kr_pcs
#  40     AN      39        kr_actual
#  44     AR      43        cn_etd
#  49     AW      48        cn_pcs        (CN ETD TOTAL)
#  50     AX      49        cn_atd
#  56     BD      55        remark        ← new position
#  57     BE      56        co            ← moved from AW
#  60     BH      59        courier
#  64     BL      63        cn_invoice_rmb ← moved from BD
#  65     BM      64        delay_reason
#  66     BN      65        cn_rdd        ← moved from BF
#  70     BR      69        fabric_etd    ← 원단 ETD (라인업 AU연동)

COL = {
    'sku':            2,
    'sort':           3,
    'class_':         4,
    'item_type':      5,
    'style_code':     7,
    'color':          8,
    'description':    9,
    'go_drop':        11,
    'designer':       14,
    'sourcing':       15,
    'vendor':         16,
    'cn_show_s':      19,
    'cn_show_l':      20,
    'cn_show_xl':     21,
    'cn_show_pcs':    22,
    'cn_focus_s':     23,
    'cn_focus_m':     24,
    'cn_focus_xl':    25,
    'cn_focus_pcs':   26,
    'kr_keep_pcs':    29,
    'ttl_pcs':        34,
    'kr_expected':    35,
    'kr_pcs':         38,
    'kr_actual':      39,
    'cn_etd':         43,
    'cn_pcs':         48,
    'cn_atd':         49,
    'remark':         55,
    'co':             56,
    'courier':        59,
    'cn_invoice_rmb': 63,
    'delay_reason':   64,
    'cn_rdd':         65,
    'fabric_etd':     69,
}

# ─── Helpers ──────────────────────────────────────────────────────────────────
def _raw(row, key):
    idx = COL[key]
    return row[idx] if idx < len(row) else None

def safe_int(v):
    if v is None:
        return 0
    if isinstance(v, str):
        s = v.strip()
        if not s or s.startswith('#'):
            return 0
        try:
            return int(float(s))
        except Exception:
            return 0
    try:
        return int(v)
    except Exception:
        return 0

def safe_float(v):
    if v is None:
        return 0.0
    if isinstance(v, str):
        s = v.strip()
        if not s or s.startswith('#'):
            return 0.0
        try:
            return float(s)
        except Exception:
            return 0.0
    try:
        return float(v)
    except Exception:
        return 0.0

def fmt_date(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        yr = v.year if isinstance(v, datetime) else v.year
        if yr < 2020:
            return None
        return v.strftime('%Y-%m-%d')
    return None

def safe_str(v):
    if v is None:
        return ''
    # Numeric 0 / 0.0 means the cell is empty (formula result)
    if isinstance(v, (int, float)) and v == 0:
        return ''
    s = str(v).strip()
    if not s or s.startswith('#') or s == '0':
        return ''
    return s

import re as _re

def fmt_fabric_etd(v):
    """원단 ETD 정제:
    - datetime/date    → YYYY-MM-DD
    - '04월 30일' 등   → 2026-MM-DD
    - 'M/D' 단순 날짜  → 2026-MM-DD
    - 'M/D~M/D' 기간  → 가장 늦은 날짜
    - '00:00:00'       → ''
    - '없음', 'TBA'    → ''
    - 'COCRE'          → 'COCRE'
    - 기타 변환 불가   → '' + 로그
    """
    if v is None:
        return ''
    if isinstance(v, (datetime, date)):
        yr = v.year
        if yr < 2020:
            return ''
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    if not s or s in ('0',) or s.startswith('#'):
        return ''
    # 이미 YYYY-MM-DD 형식
    if _re.fullmatch(r'\d{4}-\d{2}-\d{2}', s):
        return s
    # 00:00:00 (시간만)
    if _re.fullmatch(r'\d{2}:\d{2}:\d{2}', s):
        return ''
    # 미정 키워드
    if s.upper() in ('없음', 'TBA', 'TBD', 'N/A', 'NA', '-', '미정', ''):
        return ''
    # COCRE 유지
    if s.upper() == 'COCRE':
        return 'COCRE'
    # 'N월 D일' 형식 (예: '04월 30일', '4월 30일')
    m = _re.fullmatch(r'(\d{1,2})월\s*(\d{1,2})일', s)
    if m:
        return f"2026-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
    # 'M/D~M/D' 기간 → 가장 늦은 날짜
    m = _re.fullmatch(r'(\d{1,2})/(\d{1,2})\s*[~\-]\s*(\d{1,2})/(\d{1,2})', s)
    if m:
        d1 = date(2026, int(m.group(1)), int(m.group(2)))
        d2 = date(2026, int(m.group(3)), int(m.group(4)))
        return max(d1, d2).strftime('%Y-%m-%d')
    # 'M/D' 단순 형식 (예: '4/20', '5/8', '6/5')
    m = _re.fullmatch(r'(\d{1,2})/(\d{1,2})', s)
    if m:
        try:
            return date(2026, int(m.group(1)), int(m.group(2))).strftime('%Y-%m-%d')
        except ValueError:
            pass
    # 변환 실패
    print(f"  ⚠ fabric_etd 변환 실패: {repr(s)}")
    return ''

def _has_korean(s):
    return any('가' <= c <= '힣' for c in s)

# ─── Main ─────────────────────────────────────────────────────────────────────
def load_wb():
    try:
        return openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    except PermissionError:
        print(f"⚠  Excel 잠김 → 임시 복사 사용: {TMP_PATH}")
        shutil.copy2(EXCEL_PATH, TMP_PATH)
        return openpyxl.load_workbook(TMP_PATH, read_only=True, data_only=True)

def main():
    wb = load_wb()
    ws = wb['COVER']

    all_data        = []
    unmapped_vendors = set()

    for row in ws.iter_rows(min_row=4, values_only=True):
        sku = safe_str(_raw(row, 'sku') if len(row) > COL['sku'] else None)
        if not sku:
            continue

        def g(key):
            return _raw(row, key)

        # Vendor
        vendor    = safe_str(g('vendor'))
        vendor_en = VENDOR_KR_TO_EN.get(vendor)
        if vendor_en is None:
            vendor_en = vendor
            if vendor and _has_korean(vendor):
                unmapped_vendors.add(vendor)

        # CO
        co = safe_str(g('co')).upper().strip()

        # Quantities
        cn_show_pcs  = safe_int(g('cn_show_pcs'))
        cn_focus_pcs = safe_int(g('cn_focus_pcs'))
        cn_total_pcs = cn_show_pcs + cn_focus_pcs
        kr_keep_pcs  = safe_int(g('kr_keep_pcs'))
        ttl_pcs      = safe_int(g('ttl_pcs'))
        kr_pcs       = safe_int(g('kr_pcs'))
        cn_pcs       = safe_int(g('cn_pcs'))

        # is_na: item is not participating in trade show
        go_drop = safe_str(g('go_drop')).upper()
        is_na   = go_drop in ('NA', 'N/A')

        record = {
            'sku':            sku,
            'style_code':     safe_str(g('style_code')),
            'sort':           safe_str(g('sort')),
            'class_':         safe_str(g('class_')),
            'item_type':      safe_str(g('item_type')),
            'description':    safe_str(g('description')),
            'color':          safe_str(g('color')),
            'designer':       safe_str(g('designer')),
            'sourcing':       safe_str(g('sourcing')),
            'vendor':         vendor,
            'vendor_en':      vendor_en,
            'co':             co,
            'cn_invoice_rmb': round(safe_float(g('cn_invoice_rmb'))),
            'ttl_pcs':        ttl_pcs,
            'kr_pcs':         kr_pcs,
            'cn_pcs':         cn_pcs,
            'cn_show_s':      safe_int(g('cn_show_s')),
            'cn_show_l':      safe_int(g('cn_show_l')),
            'cn_show_xl':     safe_int(g('cn_show_xl')),
            'cn_show_pcs':    cn_show_pcs,
            'cn_focus_s':     safe_int(g('cn_focus_s')),
            'cn_focus_m':     safe_int(g('cn_focus_m')),
            'cn_focus_xl':    safe_int(g('cn_focus_xl')),
            'cn_focus_pcs':   cn_focus_pcs,
            'kr_keep_pcs':    kr_keep_pcs,
            'cn_total_pcs':   cn_total_pcs,
            'kr_expected':    fmt_date(g('kr_expected')),
            'kr_actual':      fmt_date(g('kr_actual')),
            'cn_etd':         fmt_date(g('cn_etd')),
            'cn_atd':         fmt_date(g('cn_atd')),
            'cn_rdd':         fmt_date(g('cn_rdd')),
            'transit_days':   TRANSIT_BY_CO.get(co, 2),
            'courier':        safe_str(g('courier')),
            'remark':         safe_str(g('remark')),
            'delay_reason':   safe_str(g('delay_reason')),
            'fabric_etd':     fmt_fabric_etd(g('fabric_etd')),
            'is_na':          is_na,
        }
        all_data.append(record)

    wb.close()

    # ─── Dashboard summaries ──────────────────────────────────────────────────
    active           = [d for d in all_data if not d['is_na']]
    total_sku        = len(all_data)
    total_styles     = len({d['style_code'] for d in all_data})
    na_styles        = len({d['style_code'] for d in all_data if d['is_na']})
    total_pcs        = sum(d['ttl_pcs']       for d in all_data)
    total_kr_keep    = sum(d['kr_keep_pcs']   for d in all_data)
    total_cn_show    = sum(d['cn_show_pcs']   for d in active)
    total_cn_focus   = sum(d['cn_focus_pcs']  for d in active)
    total_cn_total   = sum(d['cn_total_pcs']  for d in active)

    # ─── 입고예정일 (kr_expected) 분포 계산 ──────────────────────────────────
    # 스타일당 가장 빠른 kr_expected 날짜 사용
    style_earliest = {}
    for d in all_data:
        sc = d['style_code']
        if d['kr_expected']:
            if sc not in style_earliest or d['kr_expected'] < style_earliest[sc]:
                style_earliest[sc] = d['kr_expected']

    def arr_cnt(limit):
        return sum(1 for v in style_earliest.values() if v <= limit)

    arr_0612 = arr_cnt('2026-06-12')
    arr_0616 = arr_cnt('2026-06-16')
    arr_0629 = arr_cnt('2026-06-29')
    arr_0713 = arr_cnt('2026-07-13')
    ts = total_styles or 1

    arrival_summary = {
        'total_styles': total_styles,
        'by_0612': arr_0612,
        'by_0616': arr_0616,
        'by_0629': arr_0629,
        'by_0713': arr_0713,
        'pct_0612': round(arr_0612 / ts * 100, 1),
        'pct_0616': round(arr_0616 / ts * 100, 1),
        'pct_0629': round(arr_0629 / ts * 100, 1),
        'pct_0713': round(arr_0713 / ts * 100, 1),
    }

    dashboard = {
        'total_sku':           total_sku,
        'total_styles':        total_styles,
        'na_styles':           na_styles,
        'total_pcs':           total_pcs,
        'total_kr_keep_pcs':   total_kr_keep,
        'total_cn_show_pcs':   total_cn_show,
        'total_cn_focus_pcs':  total_cn_focus,
        'total_cn_total_pcs':  total_cn_total,
        'arrival_summary':     arrival_summary,
        'all_data':            all_data,
    }

    js = 'const DASHBOARD = ' + json.dumps(dashboard, ensure_ascii=False, indent=2) + ';'
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        f.write(js)

    # ─── Verification output ──────────────────────────────────────────────────
    print("=" * 52)
    print("  ✅ embed_data.js 생성 완료")
    print("=" * 52)
    print(f"  총 SKU:            {total_sku:,}")
    print(f"  총 스타일:          {total_styles:,}  (NA={na_styles})")
    print()
    print(f"  kr_keep_pcs  합계: {total_kr_keep:,}")
    print(f"  cn_show_pcs  합계: {total_cn_show:,}")
    print(f"  cn_focus_pcs 합계: {total_cn_focus:,}")
    print(f"  cn_total_pcs 합계: {total_cn_total:,}")
    print(f"  ttl_pcs      합계: {total_pcs:,}")
    check = total_kr_keep + total_cn_show + total_cn_focus
    ok = "✅ OK" if check == total_pcs else f"⚠ MISMATCH (diff={check - total_pcs:+,})"
    print(f"  검산 kr_keep+cn_show+cn_focus = {check:,}  {ok}")
    print()

    co_dist = Counter(d['co'] for d in all_data)
    print("  CO 분포:")
    for key in ('CN', 'VN', 'KR'):
        cnt = co_dist.get(key, 0)
        print(f"    {key}: {cnt:,} SKU")
    others = {k: v for k, v in co_dist.items() if k not in ('CN', 'VN', 'KR', '')}
    empty  = co_dist.get('', 0)
    for k, v in sorted(others.items()):
        print(f"    {k}: {v:,} SKU")
    if empty:
        print(f"    (빈값): {empty:,} SKU")
    print()

    total_rmb = sum(d['cn_invoice_rmb'] for d in all_data)
    print(f"  cn_invoice_rmb 총합: {total_rmb:,.0f} RMB")
    print()

    kr_exp  = sum(1 for d in all_data if d['kr_expected'])
    cn_etd_ = sum(1 for d in all_data if d['cn_etd'])
    remark_ = sum(1 for d in all_data if d['remark'])
    print(f"  kr_expected 채워진 SKU: {kr_exp:,}")
    print(f"  cn_etd      채워진 SKU: {cn_etd_:,}")
    print(f"  remark      채워진 SKU: {remark_:,}")
    print()

    # 입고예정일 분포 (스타일 기준)
    arr = arrival_summary
    print("  ★ 입고예정일 분포 (스타일 기준, 누적):")
    print(f"    6/12 이전:  {arr['by_0612']:>4} 스타일  ({arr['pct_0612']}%)  목표 40%")
    print(f"    6/16 이전:  {arr['by_0616']:>4} 스타일  ({arr['pct_0616']}%)  목표 100%")
    print(f"    6/29 이전:  {arr['by_0629']:>4} 스타일  ({arr['pct_0629']}%)  목표 60%")
    print(f"    7/13 이전:  {arr['by_0713']:>4} 스타일  ({arr['pct_0713']}%)")
    no_exp = total_styles - len(style_earliest)
    print(f"    미정(TBD):  {no_exp:>4} 스타일")
    print()

    if unmapped_vendors:
        print(f"  ⚠  미매핑 벤더 ({len(unmapped_vendors)}개):")
        for v in sorted(unmapped_vendors):
            print(f"     - {v}")
    else:
        print("  ✅ 미매핑 벤더 없음")
    print("=" * 52)

if __name__ == '__main__':
    main()
